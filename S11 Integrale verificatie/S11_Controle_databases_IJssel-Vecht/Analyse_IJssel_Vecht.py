#%%
import pandas as pd
import os
import sqlite3
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from openpyxl import load_workbook

#%%

path = {
    'Hydra_res': r'l:\C07021\C07021.000031_Controle_databases_IJssel-Vecht\Work\01_HydraNL\03_resultaten',
    'Riskeer_res': r'l:\C07021\C07021.000031_Controle_databases_IJssel-Vecht\Work\02_Riskeer\05_resultaten',
    'Riskeer_project': r'l:\C07021\C07021.000031_Controle_databases_IJssel-Vecht\Work\02_Riskeer\04_projectbestanden',
    'Analyse': r'l:\C07021\C07021.000031_Controle_databases_IJssel-Vecht\Work\03_Analyse\batch02',
    'Database': r'l:\C07021\C07021.000031_Controle_databases_IJssel-Vecht\databases'
}

parameters = ['WS', 'HS', 'HBN']

models = ['Hydra','Riskeer']

# databases =['9-2', '8-4', '7-1', '202', '10-3', '11-1', '227', '10-2', '11-2', '225']
databases =['10-2']

#%% open riskeer database and export results
def readLocationsFromHRD(path, database, onlynames=True):
    # database = 
    fn =[fn for fn in os.listdir(os.path.join(path['Database'],database)) if len(fn) >= 22 and  fn[-22:] == '_terBeoordeling.sqlite']

    if len(fn) == 1:
        HRD = os.path.join(path['Database'], database,'WBI2023_*_' + database + '_v*_terBeoordeling.sqlite')
        conn = sqlite3.connect(os.path.join(path['Database'], database, fn[0]))
        if onlynames:
            sql = '''   SELECT Name as 'Locatie'
                        FROM HRDLocations
                        ORDER BY HRDLocationId
                        '''
        else:
            sql = '''   SELECT Name as 'Locatie', XCoordinate as 'X', YCoordinate as 'Y', BedLevel as 'd'
                        FROM HRDLocations
                        ORDER BY HRDLocationId
                        '''

        # Risk = os.path.join(path['Database'], 'traject_' + database + '_HBN_10000.risk')
        # conn = sqlite3.connect(Risk)
        # if onlynames:
        #     sql = '''   SELECT Name as 'Locatie'
        #                 FROM HydraulicLocationEntity
        #                 ORDER BY HydraulicLocationEntityId
        #                 '''
        # else:
        #     sql = '''   SELECT Name as 'Locatie', LocationX as 'X', LocationY as 'Y', BedLevel as 'd'
        #                 FROM HydraulicLocationEntity
        #                 ORDER BY HydraulicLocationEntityId
        #                 '''
        df = pd.read_sql('%s' % sql, conn)

    return df

# print(readLocationsFromHRD(path, database, onlynames=True))

#%% open Riskeer result file and export results
def readHSWSFromRiskeer(conn, database, parameter, rp):
    try:
        if parameter == 'WS' and rp == 1000:
            collection = 'HydraulicLocationCalculationCollectionEntity3Id'
        elif parameter == 'WS' and rp == 10000:
            collection = 'HydraulicLocationCalculationCollectionEntity2Id'
        elif parameter == 'WS' and rp == 100000:
            collection = 'HydraulicLocationCalculationCollectionEntity2Id'
        elif parameter == 'HS' and rp == 1000:
            collection = 'HydraulicLocationCalculationCollectionEntity7Id'
        elif parameter == 'HS' and rp == 10000:
            collection = 'HydraulicLocationCalculationCollectionEntity6Id'
        elif parameter == 'HS' and rp == 100000:
            collection = 'HydraulicLocationCalculationCollectionEntity6Id'

        sql = '''
            SELECT HydraulicBoundaryDatabaseEntity.FilePath as 'Randvoorwaardendatabase', HydraulicLocationEntity.Name as 'Locatie', LocationX as 'X-coördinaat', LocationY as 'Y-coördinaat', HydraulicBoundaryDatabaseEntity.HydraulicLocationConfigurationSettingsScenarioName as 'Klimaatscenario', TargetProbability as 'Terugkeertijd [jaar]', HydraulicLocationOutputEntity.Result as 'Belastingniveau [m+NAP]/Golfparameter [m]/[s]/Sterkte bekleding [-]'
            FROM 
            AssessmentSectionEntity
            INNER JOIN HydraulicLocationCalculationEntity
            ON AssessmentSectionEntity.{0} = HydraulicLocationCalculationEntity.HydraulicLocationCalculationCollectionEntityId
            INNER JOIN HydraulicLocationOutputEntity
            ON HydraulicLocationOutputEntity.HydraulicLocationCalculationEntityId = HydraulicLocationCalculationEntity.HydraulicLocationCalculationEntityId
            INNER JOIN HydraulicLocationEntity
            ON HydraulicLocationEntity.HydraulicLocationEntityId = HydraulicLocationCalculationEntity.HydraulicLocationEntityId
            INNER JOIN HydraulicBoundaryDatabaseEntity
            ON HydraulicBoundaryDatabaseEntity.AssessmentSectionEntityId = HydraulicLocationEntity.AssessmentSectionEntityId
            WHERE AssessmentSectionEntity.Name = 'Traject {1}' AND TargetProbability = {2:f}
            ORDER BY HydraulicLocationEntity.HydraulicLocationEntityId
        '''.format(collection,database,1/rp)
        df = pd.read_sql('%s' % sql, conn)

        _, Randvoorwaardendatabase = os.path.split(df['Randvoorwaardendatabase'][0])
        df['Randvoorwaardendatabase'].replace(df['Randvoorwaardendatabase'][0],Randvoorwaardendatabase, inplace=True)
        df.insert(4, 'Profiel', '-')
        df.insert(6, 'Type Berekening', parameter)
        df.insert(7, 'Overslagdebiet [l/s/m]/Type bekleding', '-')
        df.insert(8, 'Waterstandsniveau [m+NAP]', '-')
        df['Terugkeertijd [jaar]'] = df['Terugkeertijd [jaar]'].apply(lambda x: 1/x)
        df.insert(11, 'Golfhoogte [m]', '-')
        df.insert(12, 'Piekperiode [s]', '-')
        df.insert(13, 'Golfrichting [°]', '-')
        df.insert(14, 'Golfinval [°]', '-')

    except:
        raise Exception('Failed to load sql for database %s, parameter %s for return period %s' % (database,parameter, rp))

    return df

def readHBNFromRiskeer(conn, database, parameter, rp):
    try:
        # dit statement kan compacter en completer, maar dat is voor later
        sql = '''SELECT HydraulicLocationEntity.Name as 'Locatie', LocationX as 'X-coördinaat', LocationY as 'Y-coördinaat', DikeProfileEntity.Name as 'Profiel', CriticalFlowRateMean as 'Overslagdebiet [l/s/m]/Type bekleding', TargetProbability as 'Terugkeertijd [jaar]', GrassCoverErosionInwardsDikeHeightOutputEntity.DikeHeight as 'Belastingniveau [m+NAP]/Golfparameter [m]/[s]/Sterkte bekleding [-]'
                    FROM ((((GrassCoverErosionInwardsDikeHeightOutputEntity
                    INNER JOIN GrassCoverErosionInwardsOutputEntity
                    ON GrassCoverErosionInwardsOutputEntity.GrassCoverErosionInwardsOutputEntityId = GrassCoverErosionInwardsDikeHeightOutputEntity.GrassCoverErosionInwardsOutputEntityId)
                    INNER JOIN GrassCoverErosionInwardsCalculationEntity
                    ON GrassCoverErosionInwardsCalculationEntity.GrassCoverErosionInwardsCalculationEntityId = GrassCoverErosionInwardsOutputEntity.GrassCoverErosionInwardsCalculationEntityId)
                    INNER JOIN HydraulicLocationEntity
                    ON HydraulicLocationEntity.HydraulicLocationEntityId = GrassCoverErosionInwardsCalculationEntity.HydraulicLocationEntityId)
                    INNER JOIN DikeProfileEntity
                    ON DikeProfileEntity.DikeProfileEntityId = GrassCoverErosionInwardsCalculationEntity.DikeProfileEntityId)
                    WHERE TargetProbability = {0:f}
                    ORDER BY HydraulicLocationEntity.HydraulicLocationEntityId
                    '''.format(1/rp)
        df = pd.read_sql('%s' % sql, conn)

        sql = '''SELECT HydraulicLocationConfigurationSettingsScenarioName as 'Klimaatscenario', FilePath as 'Randvoorwaardendatabase'
                    FROM HydraulicBoundaryDatabaseEntity
                    '''
        tmp = pd.read_sql('%s' % sql, conn)
        _, Randvoorwaardendatabase = os.path.split(tmp['Randvoorwaardendatabase'][0])
        df.insert(0,'Randvoorwaardendatabase',Randvoorwaardendatabase)
        df['Overslagdebiet [l/s/m]/Type bekleding'] = df['Overslagdebiet [l/s/m]/Type bekleding'].apply(lambda x: x*1000)
        df['Terugkeertijd [jaar]'] = df['Terugkeertijd [jaar]'].apply(lambda x: 1/x)
        df.insert(5, 'Klimaatscenario', tmp['Klimaatscenario'][0])
        df.insert(6, 'Type Berekening', parameter)
        df.insert(8, 'Waterstandsniveau [m+NAP]', '-')
        df.insert(11, 'Golfhoogte [m]', '-')
        df.insert(12, 'Piekperiode [s]', '-')
        df.insert(13, 'Golfrichting [°]', '-')
        df.insert(14, 'Golfinval [°]', '-')

    except:
        raise Exception('Failed to load sql for database %s, parameter %s for return period %s' % (database,parameter, rp))

    return df
 
def readResultsFromRiskeer(path, database, parameter='HBN',rp=10000):
    print('Load Riskeer',database,parameter,rp)

    # Read .risk file and extract contents
    if parameter == 'HBN':
        Risk = os.path.join(path['Riskeer_project'], 'traject_' + database + '_HBN_10000.risk')
    else:
        if rp == 100000:
            if database == '10-2':
                Risk = os.path.join(path['Riskeer_project'], 'traject_10-2_freq_100000.risk')
            elif database in ['225','227','8-4','9-2']:
                Risk = os.path.join(path['Riskeer_project'], 'traject_225_227_8-4_9-2_freq_100000.risk')
            else:
                Risk = os.path.join(path['Riskeer_project'], 'traject_10-3_11-1_11-2_202_7-1_freq_100000.risk')
        else:
            if database == '10-2':
                Risk = os.path.join(path['Riskeer_project'], 'traject_10-2_freq_1000_10000.risk')
            elif database in ['225','227','8-4','9-2']:
                Risk = os.path.join(path['Riskeer_project'], 'traject_225_227_8-4_9-2_freq_1000_10000.risk')
            else:
                Risk = os.path.join(path['Riskeer_project'], 'traject_10-3_11-1_11-2_202_7-1_freq_1000_10000.risk')

    conn = sqlite3.connect(Risk)

    try:
        if parameter =='WS' or parameter == 'HS':
            df = readHSWSFromRiskeer(conn, database, parameter,rp)
        elif parameter == 'HBN':
            df = readHBNFromRiskeer(conn, database, parameter, rp)
        else:
            raise Exception('Failed to determine results for database %s, parameter %s' % (database,parameter))
    except:
        raise Exception('Failed to determine results for database %s, parameter %s and return period %s' % (database,parameter,rp))

    Locaties = readLocationsFromHRD(path, database)
    df = pd.merge(df,Locaties,how='outer', on=['Locatie']).sort_values('Locatie')

    excelpath = os.path.join(path['Riskeer_res'], parameter + '_' + database + '_' + '%s' % (rp) + '.xls')
    if not os.path.exists(excelpath): 
        df.to_excel(excelpath, sheet_name=parameter + '_' + database, index=False, encoding='ISO-8859-1')

    return df
# Riskeer = readResultsFromRiskeer(path, database='225', parameter='WS', rp=1000)

#%% open Hydra result file and export results
def readResultsFromHydra(path, database, parameter='HBN',rp=10000):
    print('Load Hydra',database,parameter,rp)
    hydraResFile = os.path.join(path['Hydra_res'], parameter + '_' + database + '.xls')
    df = pd.read_csv(hydraResFile, sep='\t', header=0, encoding = "ISO-8859-1")
    df = df[df['Terugkeertijd [jaar]'] == rp]

    Locaties = readLocationsFromHRD(path, database)
    df = pd.merge(df,Locaties,how='outer', on=['Locatie']).sort_values('Locatie')

    return df
# Hydra = readResultsFromHydra(path, database='10-2', parameter='HS', rp=10000)

#%% analyse van de resultaten
def analyseGecombineerd(Gecombineerd):

    # analyse combined results
    for rp in [1000,10000,100000]:
        Gecombineerd['Hydra_dWS_{:d}'.format(rp)] = Gecombineerd['Hydra_WS_{:d}'.format(rp)].diff()
        Gecombineerd['Riskeer_dWS_{:d}'.format(rp)] = Gecombineerd['Riskeer_WS_{:d}'.format(rp)].diff()
        Gecombineerd['Hydra_dHS_{:d}'.format(rp)] = Gecombineerd['Hydra_HS_{:d}'.format(rp)].diff()
        Gecombineerd['Riskeer_dHS_{:d}'.format(rp)] = Gecombineerd['Riskeer_HS_{:d}'.format(rp)].diff()

        Gecombineerd['dWS_{:d}'.format(rp)] = Gecombineerd['Hydra_WS_{:d}'.format(rp)] - Gecombineerd['Riskeer_WS_{:d}'.format(rp)]
        Gecombineerd['dHS_{:d}'.format(rp)] = Gecombineerd['Hydra_HS_{:d}'.format(rp)] - Gecombineerd['Riskeer_HS_{:d}'.format(rp)]
    Gecombineerd['dHBN_10000'] = Gecombineerd['Hydra_HBN_10000'] - Gecombineerd['Riskeer_HBN_10000']

    Gecombineerd['Hydra_dchWS_1000'] = Gecombineerd['Hydra_WS_10000'] - Gecombineerd['Hydra_WS_1000']
    Gecombineerd['Hydra_dchWS_100000'] = Gecombineerd['Hydra_WS_100000'] - Gecombineerd['Hydra_WS_10000']
    Gecombineerd['Hydra_dchHS_1000'] = Gecombineerd['Hydra_HS_10000'] - Gecombineerd['Hydra_HS_1000']
    Gecombineerd['Hydra_dchHS_100000'] = Gecombineerd['Hydra_HS_100000'] - Gecombineerd['Hydra_HS_10000']

    Gecombineerd['Riskeer_dchWS_1000'] = Gecombineerd['Riskeer_WS_10000'] - Gecombineerd['Riskeer_WS_1000']
    Gecombineerd['Riskeer_dchWS_100000'] = Gecombineerd['Riskeer_WS_100000'] - Gecombineerd['Riskeer_WS_10000']
    Gecombineerd['Riskeer_dchHS_1000'] = Gecombineerd['Riskeer_HS_10000'] - Gecombineerd['Riskeer_HS_1000']
    Gecombineerd['Riskeer_dchHS_100000'] = Gecombineerd['Riskeer_HS_100000'] - Gecombineerd['Riskeer_HS_10000']

    Gecombineerd['ddch_WS_1000'] = Gecombineerd['Hydra_dchWS_1000'] - Gecombineerd['Riskeer_dchWS_1000']
    Gecombineerd['ddch_WS_100000'] = Gecombineerd['Hydra_dchWS_100000'] - Gecombineerd['Riskeer_dchWS_100000']
    Gecombineerd['ddch_HS_1000'] = Gecombineerd['Hydra_dchHS_1000'] - Gecombineerd['Riskeer_dchHS_1000']
    Gecombineerd['ddch_HS_100000'] = Gecombineerd['Hydra_dchHS_100000'] - Gecombineerd['Riskeer_dchHS_100000']
    # Gecombineerd['ddch_HBN_1000'] = Gecombineerd['Hydra_dchHBN_1000'] - Gecombineerd['Riskeer_dchHBN_1000']
    # Gecombineerd['ddch_HBN_100000'] = Gecombineerd['Hydra_dchHBN_100000'] - Gecombineerd['Riskeer_dchHBN_100000']

    empty = pd.DataFrame([[np.nan] * len(Gecombineerd.columns)], columns=Gecombineerd.columns)
    Gecombineerd = empty.append(Gecombineerd, ignore_index=True)
    # for item in ['L','S','Hydra_dWS_10000','Riskeer_dWS_10000','dWS_10000','dHS_10000','dHBN_10000','ddch_WS_1000','ddch_WS_100000']:
    for i in range(3,30):
        max = Gecombineerd.iloc[:,i].max()
        Gecombineerd.at[0, Gecombineerd.columns[i]] = max
        Gecombineerd.sort_index(inplace=True)
        print('Max ' + Gecombineerd.columns[i],database,parameter,returnPeriod,'{:.1f}'.format(max))

    return Gecombineerd

#%% afbeelding van de resultaten
def plotResultsGecombineerd(path, database, Gecombineerd):

    plotResults(path, database, Gecombineerd,'WS')
    plotResults(path, database, Gecombineerd,'HS')
    plotResults(path, database, Gecombineerd,'HBN')
    plotDifferences(path, database, Gecombineerd)
    plotDecimeringshoogte(path, database, Gecombineerd)

def plotResults(path, database, Gecombineerd, parameter='HBN'):

    label ={'WS': 'Waterstand','HS': 'Golfhoogte','HBN': 'Hydraulisch belastingniveau'}
    units ={'WS': '[m+NAP]','HS': '[m]','HBN': '[m]'}

    plt.figure(1,figsize=(6, 3.5))
    fig, ax = plt.subplots(nrows=1, ncols=1)
    cm1 = plt.cm.Blues(np.linspace(0,1,4))
    cm2 = plt.cm.Reds(np.linspace(0,1,4))

    if not parameter == 'HBN':
        plt.plot(Gecombineerd['S']/1000,Gecombineerd['Hydra_' + parameter + '_1000'], label='    Hydra - Terugkeerperiode 1000 [jaar]', color=cm1[1,:])
        plt.plot(Gecombineerd['S']/1000,Gecombineerd['Hydra_' + parameter + '_10000'], label='    Hydra - Terugkeerperiode 10000 [jaar]', color=cm1[2,:])
        plt.plot(Gecombineerd['S']/1000,Gecombineerd['Hydra_' + parameter + '_100000'], label='    Hydra - Terugkeerperiode 100000 [jaar]', color=cm1[3,:])
        plt.plot(Gecombineerd['S']/1000,Gecombineerd['Riskeer_' + parameter + '_1000'], label='    Riskeer - Terugkeerperiode 1000 [jaar]', color=cm2[1,:])
        plt.plot(Gecombineerd['S']/1000,Gecombineerd['Riskeer_' + parameter + '_10000'], label='    Riskeer - Terugkeerperiode 10000 [jaar]', color=cm2[2,:])
        plt.plot(Gecombineerd['S']/1000,Gecombineerd['Riskeer_' + parameter + '_100000'], label='    Riskeer - Terugkeerperiode 100000 [jaar]', color=cm2[3,:])
    else:
        plt.plot(Gecombineerd['S']/1000,Gecombineerd['Hydra_' + parameter + '_10000'], label='    Hydra - Terugkeerperiode  10000 [jaar]', color=cm1[2,:])
        plt.plot(Gecombineerd['S']/1000,Gecombineerd['Riskeer_' + parameter + '_10000'], label='    Riskeer - Terugkeerperiode  10000 [jaar]', color=cm2[2,:])

    plt.xlabel('Afstand tot eerste steunpuntlocatie [km]')
    plt.ylabel(label[parameter] + ' ' + units[parameter])
    # plt.xlim(0, 15)
    # plt.ylim(0, 10)
    y_labels = ax.get_yticks()
    ax.yaxis.set_major_formatter(ticker.FormatStrFormatter('%.2f'))

    plt.grid()
    plt.legend(ncol=2, loc='lower left', 
            bbox_to_anchor=[0, 1], 
            columnspacing=1.0, labelspacing=0.0,
            handletextpad=0.0, handlelength=1.5,
            frameon=False, shadow=False, fontsize=6)
    plt.savefig(os.path.join(path['Analyse'], 'Results_' + database + '_' + parameter), dpi=600, facecolor='w')
    plt.close()
    
def plotDifferences(path, database, Gecombineerd, rp=10000):

    labels = {'WS': 'Waterstand','HS': 'Golfhoogte','HBN': 'Hydraulisch belastingniveau'}
    units = {'WS': '[m+NAP]','HS': '[m]','HBN': '[m]'}

    plt.figure(1,figsize=(6, 3.5))
    fig, ax = plt.subplots(nrows=1, ncols=1)
    cm1 = plt.cm.Blues(np.linspace(0,1,4))

    for i, parameter in enumerate(['WS','HS','HBN']):
        plt.plot(Gecombineerd['S']/1000,Gecombineerd['d{}_{:d}'.format(parameter,rp)], label='    Verschil {} Hydra-Riskeer - Terugkeerperiode {:d} [jaar]'.format(labels[parameter], rp), color=cm1[i+1,:])

    plt.xlabel('Afstand tot eerste steunpuntlocatie [km]')
    plt.ylabel('Verschil [m]')
    # plt.xlim(0, 15)
    # plt.ylim(0, 10)
    y_labels = ax.get_yticks()
    ax.yaxis.set_major_formatter(ticker.FormatStrFormatter('%.2f'))
    plt.grid()
    plt.legend(ncol=1, loc='lower left', 
            bbox_to_anchor=[0, 1], 
            columnspacing=1.0, labelspacing=0.0,
            handletextpad=0.0, handlelength=1.5,
            frameon=False, shadow=False, fontsize=6)
    plt.savefig(os.path.join(path['Analyse'], 'Difference_{}_{:d}'.format(database,rp)), dpi=600, facecolor='w')
    plt.close()

def plotDecimeringshoogte(path, database, Gecombineerd):

    label ={'WS': 'Waterstand','HS': 'Golfhoogte','HBN': 'Hydraulisch belastingniveau'}
    units ={'WS': '[m+NAP]','HS': '[m]','HBN': '[m]'}

    plt.figure(1,figsize=(6, 3.5))
    fig, ax = plt.subplots(nrows=1, ncols=1)
    cm1 = plt.cm.Blues(np.linspace(0,1,4))
    cm2 = plt.cm.Reds(np.linspace(0,1,4))


    plt.plot(Gecombineerd['S']/1000,Gecombineerd['Hydra_dchWS_1000'], label='    Hydra - Decimeringshoogte ' + label['WS'] +  ' Terugkeerperiode 10000  - 1000 [jaar]', color=cm1[1,:], linewidth=2)
    plt.plot(Gecombineerd['S']/1000,Gecombineerd['Hydra_dchWS_100000'], label='    Hydra - Decimeringshoogte ' + label['WS'] +  ' Terugkeerperiode 100000  - 10000 [jaar]', color=cm1[3,:], linewidth=2)
    plt.plot(Gecombineerd['S']/1000,Gecombineerd['Riskeer_dchWS_1000'], label='    Riskeer - Decimeringshoogte ' + label['WS'] +  ' Terugkeerperiode 10000  - 1000 [jaar]', color=cm2[1,:], linewidth=2)
    plt.plot(Gecombineerd['S']/1000,Gecombineerd['Riskeer_dchWS_100000'], label='    Riskeer - Decimeringshoogte ' + label['WS'] +  ' Terugkeerperiode 100000  - 10000 [jaar]', color=cm2[3,:], linewidth=2)
    

    plt.xlabel('Afstand tot eerste steunpuntlocatie [km]')
    plt.ylabel('Decimeringshoogte [m]')
    # plt.xlim(0, 15)
    # plt.ylim(0, 10)
    y_labels = ax.get_yticks()
    ax.yaxis.set_major_formatter(ticker.FormatStrFormatter('%.2f'))
    plt.grid()
    plt.legend(ncol=1, loc='lower left', 
            bbox_to_anchor=[0, 1], 
            columnspacing=1.0, labelspacing=0.0,
            handletextpad=0.0, handlelength=1.5,
            frameon=False, shadow=False, fontsize=6)
    plt.savefig(os.path.join(path['Analyse'], 'Decimeringshoogte_' + database), dpi=600, facecolor='w')
    plt.close()

# %%
# check ontbrekende waardes en droogval (waarde = 0.0)
for database in databases:
    # create and populate Gecombineerd
    Gecombineerd = readLocationsFromHRD(path, database, onlynames=False)
    Gecombineerd['L'] = (Gecombineerd['X'].diff().pow(2, axis=0) + Gecombineerd['Y'].diff().pow(2, axis=0)).pow(0.5, axis=0)
    Gecombineerd['S'] = Gecombineerd['L'].cumsum()

    # add results for various models, parameters and return periods
    for model in models:
        for parameter in parameters:
            for returnPeriod in [1000, 10000, 100000]:
                if parameter == 'HBN' and not returnPeriod == 10000:
                    continue
                
                if model == 'Hydra':
                    Resultaat = readResultsFromHydra(path, database, parameter=parameter, rp=returnPeriod)
                else:
                    Resultaat = readResultsFromRiskeer(path, database, parameter=parameter, rp=returnPeriod)

                Gecombineerd = Gecombineerd.merge(Resultaat[['Locatie','Belastingniveau [m+NAP]/Golfparameter [m]/[s]/Sterkte bekleding [-]']], on='Locatie').rename(columns={'Belastingniveau [m+NAP]/Golfparameter [m]/[s]/Sterkte bekleding [-]': model + '_' + parameter + '_' + str(returnPeriod)}, inplace=False)
    
    Gecombineerd = analyseGecombineerd(Gecombineerd)
    plotResultsGecombineerd(path, database, Gecombineerd)
    plotDifferences(path, database, Gecombineerd, rp=100000)

    if True:
        Gecombineerd.rename(index={0: 'max'}, inplace=True)
        excelfile = os.path.join(path['Analyse'], 'Analyse_IJssel_Vecht.xlsx')
        with pd.ExcelWriter(excelfile, engine = 'openpyxl') as writer:
            if  os.path.exists(excelfile): 
                # book = load_workbook(os.path.join(path['Analyse'], 'Analyse_IJssel_Vecht.xlsx'))
                writer.book = load_workbook(excelfile)
                Gecombineerd.to_excel(writer, sheet_name=database, float_format='%.3f')
            else:
                # Gecombineerd.to_excel(excelfile, sheet_name=parameter + '_' + database, index=False, encoding='ISO-8859-1')
                Gecombineerd.to_excel(excelfile, sheet_name=database, float_format='%.3f')

# %%
